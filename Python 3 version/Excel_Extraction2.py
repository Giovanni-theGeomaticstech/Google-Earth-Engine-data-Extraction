################################################################################################################
#Author : Giovanni Harvey
#Last Modified Date: August 23, 2019
#Job Title: Google Earth Engine API Developer
#May 2019 - August 2019
################################################################################################################

import os 
from openpyxl import Workbook
import re
#import pypickle
#from pypickle import *

# Test Dictioanry 
#dict = {'Barren' : [['A',0],['B',1],['C',2],['D',3],['E',4],['F',5]],'Cropland' : [['A',0],['B',1],['C',2],['D',3],['E',4],['F',5]]}

#k = read_Object("C:/workspace/datalab-ee/datalab/notebooks/Extraction_proj/TestersDict.pickle")
######################################################################################################################

def name_reduce(key):
    assert key, "Something wrong with user key"

    index = key.find(":") #colon index
    n_key = key[0:index]

    return n_key


def extraction(dict,image,year):
    assert dict,"Your dictionary creation has errors present in it"
        #Here lets do a test excel table generation
    wb = Workbook()
    tmp_shts = wb.sheetnames
    #get year for extraction
    #use to check in to see if the year in present in the id
    wb[tmp_shts[0]].title = 'Extraction'
    tmp_shts = wb.sheetnames
    ws = wb[tmp_shts[0]]
    

    ##Buillding the worksheet 

    ws.cell(row=1,column=1).value = 'Station_Names'

    i = 2 #column for changing the columns
    for key in dict: #all the landcover/extracted info the dictionary
        temp_key = name_reduce(key)

        ws.cell(row=1,column=i).value = temp_key #putting the keys into
        ws.cell(row=1,column=i+1).value = '% Area of ' + temp_key #adding the percentage column to twith the key associated
        
        k = 2 #rows for the changing rows 
        for stat_list in dict[key]: #all the elements present in dictionary at that info
            for data in stat_list:
                station = ws.cell(row=k,column=1).value
                if (station == None): #Populating station data into excel cell
                    ws.cell(row=k,column=1).value = data[0]
                    ws.cell(row=k,column=i).value = data[1]
                    ws.cell(row=k,column=i+1).value = data[2]
                    k+=1
                else:
                    ws.cell(row=k,column=i).value = data[1]
                    ws.cell(row=k,column=i+1).value = data[2]
                    k+=1
        i+=2
    print("Extraction completed \n")
    print("Document will now be open!")

    if(run(image)):
        image = new_string(image)
    
    dirs = "C:\GEE_Extraction"
    if (os.path.isdir(dirs)):
        os.chdir(dirs)
    else:
        os.mkdir(dirs)
    file_name = dirs + "\\" + image + year + ".xlsx"
    wb.save(file_name) #put a fail save for sheet already created
    os.startfile(file_name)#"extraction_doc"



######################################################################################################################


    
# Python program to check if a string 
# contains any special character 
  
# import required package 

  
# Function checks if the string 
# contains any special character 
def run(string): 
    # Make own character set and pass  
    # this as argument in compile method 
    regex = re.compile('[@_!#$%^&*"()<>?/|\}{~:]') 
      
    # Pass the string in search  
    # method of regex object.     
    if(regex.search(string) == None): 
        return False
    else: 
        return True 
      
def new_string(string):
    # Make own character set and pass  
    # this as argument in compile method 
    regex = re.compile('[@_!#$%^&*()<>?/\|}{~:]')
    newValue = re.split(regex,string)
    string = '_'
    string = string.join(newValue)
    return string
        
#extraction(k,"testeers")
    



